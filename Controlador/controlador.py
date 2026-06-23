from flask import Flask, render_template, request, redirect, url_for, session, jsonify
import sys
import os
import ssl
import time
import random
import hashlib
import smtplib
import pyotp
from email.mime.text import MIMEText
from datetime import datetime, timedelta

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'Modelo'))
import modelo

app = Flask(__name__, template_folder='../Vista', static_folder='../static')
app.secret_key = 'taller_el_costa_2026'


# Formatea una fecha (date o texto 'YYYY-MM-DD') como "23 jun 2026". Útil en las vistas.
_MESES_ES = ['ene', 'feb', 'mar', 'abr', 'may', 'jun', 'jul', 'ago', 'sep', 'oct', 'nov', 'dic']

@app.template_filter('fecha_corta')
def fecha_corta(v):
    if not v:
        return '—'
    try:
        if isinstance(v, str):
            v = datetime.strptime(v[:10], '%Y-%m-%d').date()
        return f"{v.day:02d} {_MESES_ES[v.month - 1]} {v.year}"
    except Exception:
        return str(v)


# ═════════════════════════════════════════
#  LOGIN
# ═════════════════════════════════════════

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        correo     = request.form.get('correo')
        contrasena = request.form.get('contrasena')
        usuario    = modelo.login_usuario(correo, contrasena)
        if usuario:
            if usuario.get('totp_secret'):
                # Tiene 2FA activo: pedir el código de la app antes de entrar
                session['pre2fa_id'] = usuario['idUsuario']
                return redirect(url_for('login_2fa'))
            iniciar_sesion(usuario)
            return redirect(url_for('dashboard'))
        else:
            return render_template('login.html', error='Correo o contraseña incorrectos', correo=correo)
    return render_template('login.html')


def iniciar_sesion(usuario):
    """Marca la sesión como autenticada con los datos del usuario."""
    session.pop('pre2fa_id', None)
    session['usuario']    = usuario['nombre'] + ' ' + (usuario['apellido'] or '')
    session['id_rol']     = usuario['id_Rol_fk']
    session['id_usuario'] = usuario['idUsuario']
    session['correo']     = usuario.get('correo')


@app.route('/login/2fa', methods=['GET', 'POST'])
def login_2fa():
    uid = session.get('pre2fa_id')
    if not uid:
        return redirect(url_for('login'))
    usuario = modelo.obtener_usuario_por_id(uid)
    if not usuario or not usuario.get('totp_secret'):
        return redirect(url_for('login'))
    if request.method == 'POST':
        codigo = (request.form.get('codigo') or '').strip()
        if pyotp.TOTP(usuario['totp_secret']).verify(codigo, valid_window=1):
            iniciar_sesion(usuario)
            return redirect(url_for('dashboard'))
        return render_template('login_2fa.html', error='Código incorrecto. Revisa tu app Authenticator.')
    return render_template('login_2fa.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ═════════════════════════════════════════
#  RECUPERAR CONTRASEÑA (código por correo)
# ═════════════════════════════════════════

def enviar_codigo_correo(destino, codigo):
    """Envía el código por Gmail si hay credenciales (MAIL_USER/MAIL_PASSWORD);
    si no, lo imprime en consola (modo demo). Devuelve True si se envió por correo real."""
    mail_user = os.environ.get('MAIL_USER')
    mail_pass = os.environ.get('MAIL_PASSWORD')
    cuerpo = (
        f"Hola,\n\nTu código para recuperar la contraseña de Taller El Costa es:\n\n"
        f"    {codigo}\n\nEste código vence en 10 minutos. Si no fuiste tú, ignora este correo."
    )
    if not mail_user or not mail_pass:
        print(f"[DEMO] Código de recuperación para {destino}: {codigo}")
        return False
    msg = MIMEText(cuerpo)
    msg['Subject'] = 'Código de recuperación — Taller El Costa'
    msg['From'] = mail_user
    msg['To'] = destino
    contexto = ssl.create_default_context()
    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=contexto) as servidor:
        servidor.login(mail_user, mail_pass)
        servidor.sendmail(mail_user, destino, msg.as_string())
    return True


def _hash_codigo(codigo):
    return hashlib.sha256((app.secret_key + codigo).encode()).hexdigest()


@app.route('/recuperar', methods=['GET', 'POST'])
def recuperar():
    if request.method == 'POST':
        correo = (request.form.get('correo') or '').strip()
        usuario = modelo.obtener_usuario_por_correo(correo)
        if not usuario:
            return render_template('recuperar.html', paso=1, error='No existe una cuenta con ese correo', correo=correo)
        codigo = '%06d' % random.randint(0, 999999)
        session['rec_correo'] = correo
        session['rec_hash']   = _hash_codigo(codigo)
        session['rec_exp']    = time.time() + 600  # 10 minutos
        session.pop('rec_ok', None)
        enviado = enviar_codigo_correo(correo, codigo)
        aviso = None if enviado else f'Modo demo: tu código es {codigo} (revisa la consola del servidor)'
        return render_template('recuperar.html', paso=2, correo=correo, aviso=aviso)
    return render_template('recuperar.html', paso=1)


@app.route('/recuperar/codigo', methods=['POST'])
def recuperar_codigo():
    correo = session.get('rec_correo')
    if not correo:
        return redirect(url_for('recuperar'))
    codigo = (request.form.get('codigo') or '').strip()
    if time.time() > session.get('rec_exp', 0):
        return render_template('recuperar.html', paso=1, error='El código venció. Solicítalo de nuevo.')
    if _hash_codigo(codigo) != session.get('rec_hash'):
        return render_template('recuperar.html', paso=2, correo=correo, error='Código incorrecto')
    session['rec_ok'] = correo
    return render_template('recuperar.html', paso=3, correo=correo)


@app.route('/recuperar/nueva', methods=['POST'])
def recuperar_nueva():
    correo = session.get('rec_ok')
    if not correo:
        return redirect(url_for('recuperar'))
    nueva = request.form.get('contrasena') or ''
    confirmar = request.form.get('confirmar') or ''
    if len(nueva) < 4:
        return render_template('recuperar.html', paso=3, correo=correo, error='La contraseña debe tener al menos 4 caracteres')
    if nueva != confirmar:
        return render_template('recuperar.html', paso=3, correo=correo, error='Las contraseñas no coinciden')
    modelo.actualizar_contrasena_por_correo(correo, nueva)
    for k in ('rec_correo', 'rec_hash', 'rec_exp', 'rec_ok'):
        session.pop(k, None)
    return render_template('login.html', exito='Contraseña actualizada. Ya puedes iniciar sesión.')


# ═════════════════════════════════════════
#  SEGURIDAD — 2FA (Google Authenticator)
# ═════════════════════════════════════════

ISSUER_2FA = 'Taller El Costa'


@app.route('/seguridad')
def seguridad():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    usuario = modelo.obtener_usuario_por_id(session.get('id_usuario'))
    activo = bool(usuario and usuario.get('totp_secret'))
    return render_template('seguridad.html', activo=activo)


@app.route('/seguridad/activar', methods=['POST'])
def seguridad_activar():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    usuario = modelo.obtener_usuario_por_id(session.get('id_usuario'))
    secret = pyotp.random_base32()
    session['pending_totp'] = secret
    etiqueta = (usuario.get('correo') if usuario else None) or session.get('usuario')
    uri = pyotp.totp.TOTP(secret).provisioning_uri(name=etiqueta, issuer_name=ISSUER_2FA)
    return render_template('seguridad.html', activo=False, configurando=True, secret=secret, uri=uri)


@app.route('/seguridad/confirmar', methods=['POST'])
def seguridad_confirmar():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    secret = session.get('pending_totp')
    if not secret:
        return redirect(url_for('seguridad'))
    codigo = (request.form.get('codigo') or '').strip()
    if pyotp.TOTP(secret).verify(codigo, valid_window=1):
        modelo.actualizar_totp_secret(session.get('id_usuario'), secret)
        session.pop('pending_totp', None)
        return render_template('seguridad.html', activo=True,
                               exito='¡Verificación en 2 pasos activada! Ahora se pedirá el código al iniciar sesión.')
    etiqueta = session.get('correo') or session.get('usuario')
    uri = pyotp.totp.TOTP(secret).provisioning_uri(name=etiqueta, issuer_name=ISSUER_2FA)
    return render_template('seguridad.html', activo=False, configurando=True, secret=secret, uri=uri,
                           error='Código incorrecto, inténtalo de nuevo.')


@app.route('/seguridad/desactivar', methods=['POST'])
def seguridad_desactivar():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    modelo.actualizar_totp_secret(session.get('id_usuario'), None)
    session.pop('pending_totp', None)
    return render_template('seguridad.html', activo=False, exito='Verificación en 2 pasos desactivada.')


# ═════════════════════════════════════════
#  DASHBOARD
# ═════════════════════════════════════════

MODULOS_POR_ROL = {
    1: ['empleados', 'inventario', 'ordenes', 'facturacion', 'proveedores', 'estadisticas'],  # Super_administrador (R001: total)
    2: ['empleados', 'inventario', 'ordenes', 'facturacion', 'proveedores', 'estadisticas'],  # Administrador (R002: gestiona usuarios/productos/órdenes + reportes)
    3: [],                                                                                     # Cliente (R003: sin acceso al programa)
    4: ['ordenes'],                                                                            # Mecánico (R004: solo órdenes asignadas)
}


def solo_administradores():
    # Permite Super_administrador (1) y Administrador (2). Redirige al resto.
    # R001/R002: ambos pueden crear, editar y eliminar usuarios.
    if 'usuario' not in session:
        return redirect(url_for('login'))
    if session.get('id_rol') not in (1, 2):
        return redirect(url_for('dashboard'))
    return None

@app.route('/dashboard')
def dashboard():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    permitidos = MODULOS_POR_ROL.get(session.get('id_rol'), [])
    return render_template('dashboard.html', modulos=permitidos)

# ═════════════════════════════════════════
#  USUARIOS
# ═════════════════════════════════════════

@app.route('/usuarios')
def usuarios():
    guard = solo_administradores()
    if guard: return guard
    lista = modelo.obtener_usuarios()
    roles = modelo.obtener_roles()
    return render_template('usuarios.html', usuarios=lista, roles=roles)

@app.route('/usuarios/agregar', methods=['POST'])
def agregar_usuario():
    guard = solo_administradores()
    if guard: return guard
    nombre     = request.form.get('nombre')
    apellido   = request.form.get('apellido')
    contrasena = request.form.get('contrasena')
    correo     = request.form.get('correo')
    id_rol     = request.form.get('id_rol')
    modelo.crear_usuario(nombre, apellido, contrasena, correo, id_rol)
    return redirect(url_for('usuarios'))

@app.route('/usuarios/editar/<int:id>', methods=['POST'])
def editar_usuario(id):
    guard = solo_administradores()
    if guard: return guard
    nombre     = request.form.get('nombre')
    apellido   = request.form.get('apellido')
    contrasena = request.form.get('contrasena')
    correo     = request.form.get('correo')
    id_rol     = request.form.get('id_rol')
    modelo.actualizar_usuario(id, nombre, apellido, contrasena, correo, id_rol)
    return redirect(url_for('usuarios'))

@app.route('/usuarios/eliminar/<int:id>')
def eliminar_usuario(id):
    guard = solo_administradores()
    if guard: return guard
    u = modelo.obtener_usuario_por_id(id)
    nombre = ((u.get('nombre') or '') + ' ' + (u.get('apellido') or '')).strip() if u else ''
    # No se puede borrar si tiene órdenes/atenciones asociadas (la BD lo protege con FK)
    dep = modelo.contar_dependencias_usuario(id)
    if dep['total'] > 0:
        return redirect(url_for('usuarios', err='rel', nombre=nombre, n=dep['ordenes']))
    try:
        modelo.eliminar_usuario(id)
    except Exception:
        return redirect(url_for('usuarios', err='rel', nombre=nombre))
    return redirect(url_for('usuarios', ok='del', nombre=nombre))


# ═════════════════════════════════════════
#  VEHÍCULOS
# ═════════════════════════════════════════

@app.route('/vehiculos')
def vehiculos():
    lista    = modelo.obtener_vehiculos()
    usuarios = modelo.obtener_usuarios()
    marcas   = modelo.obtener_marcas()
    return render_template('vehiculos.html', vehiculos=lista, usuarios=usuarios, marcas=marcas)

@app.route('/vehiculos/agregar', methods=['POST'])
def agregar_vehiculo():
    id_usuario = request.form.get('id_usuario')
    id_marca   = request.form.get('id_marca')
    placa      = request.form.get('placa')
    mod        = request.form.get('modelo')
    anio       = request.form.get('anio')
    color      = request.form.get('color')
    tipo       = request.form.get('tipo')
    modelo.crear_vehiculo(id_usuario, id_marca, placa, mod, anio, color, tipo)
    return redirect(url_for('vehiculos'))

@app.route('/vehiculos/editar/<int:id>', methods=['POST'])
def editar_vehiculo(id):
    id_usuario = request.form.get('id_usuario')
    id_marca   = request.form.get('id_marca')
    placa      = request.form.get('placa')
    mod        = request.form.get('modelo')
    anio       = request.form.get('anio')
    color      = request.form.get('color')
    tipo       = request.form.get('tipo')
    modelo.actualizar_vehiculo(id, id_usuario, id_marca, placa, mod, anio, color, tipo)
    return redirect(url_for('vehiculos'))

@app.route('/vehiculos/eliminar/<int:id>')
def eliminar_vehiculo(id):
    modelo.eliminar_vehiculo(id)
    return redirect(url_for('vehiculos'))


# ═════════════════════════════════════════
#  INVENTARIO
# ═════════════════════════════════════════

@app.route('/inventario')
def inventario():
    productos  = modelo.obtener_productos()
    categorias = modelo.obtener_categorias()
    return render_template('inventario.html', productos=productos, categorias=categorias)

@app.route('/inventario/agregar', methods=['POST'])
def agregar_producto():
    id_categoria  = request.form.get('id_categoria')
    nombre        = request.form.get('nombre')
    descripcion   = request.form.get('descripcion')
    stock         = request.form.get('stock', 0)
    stock_minimo  = request.form.get('stock_minimo', 0)
    precio_compra = request.form.get('precio_compra', 0)
    precio_venta  = request.form.get('precio_venta', 0)
    modelo.crear_producto(id_categoria, nombre, descripcion, stock, stock_minimo, precio_compra, precio_venta)
    return redirect(url_for('inventario'))

@app.route('/inventario/editar/<int:id>', methods=['POST'])
def editar_producto(id):
    id_categoria  = request.form.get('id_categoria')
    nombre        = request.form.get('nombre')
    descripcion   = request.form.get('descripcion')
    stock         = request.form.get('stock', 0)
    stock_minimo  = request.form.get('stock_minimo', 0)
    precio_compra = request.form.get('precio_compra', 0)
    precio_venta  = request.form.get('precio_venta', 0)
    modelo.actualizar_producto(id, id_categoria, nombre, descripcion, stock, stock_minimo, precio_compra, precio_venta)
    return redirect(url_for('inventario'))

@app.route('/inventario/eliminar/<int:id>')
def eliminar_producto(id):
    modelo.eliminar_producto(id)
    return redirect(url_for('inventario'))


# ═════════════════════════════════════════
#  ÓRDENES DE SERVICIO + ATENCIONES
# ═════════════════════════════════════════

@app.route('/ordenes')
def ordenes():
    # El mecánico (rol 4) solo ve SUS órdenes y SUS atenciones; admin/super ven todo.
    if session.get('id_rol') == 4:
        uid        = session.get('id_usuario')
        lista      = modelo.obtener_ordenes_por_mecanico(uid)
        atenciones = modelo.obtener_atenciones_por_mecanico(uid)
    else:
        lista      = modelo.obtener_ordenes()
        atenciones = modelo.obtener_atenciones()
    lista_veh  = modelo.obtener_vehiculos()
    lista_usu  = modelo.obtener_usuarios()
    estados    = modelo.obtener_estados_orden()
    siguiente  = modelo.obtener_siguiente_numero_orden()
    marcas     = modelo.obtener_marcas()
    return render_template('Orden_de_servicio.html',
                           ordenes=lista,
                           vehiculos=lista_veh,
                           usuarios=lista_usu,
                           estados=estados,
                           atenciones=atenciones,
                           siguiente_orden=siguiente,
                           marcas=marcas,
                           productos=modelo.obtener_productos(),
                           tipos_servicio=modelo.obtener_tipos_servicio(),
                           id_rol=session.get('id_rol'))

@app.route('/ordenes/agregar', methods=['POST'])
def agregar_orden():
    id_mecanico  = request.form.get('id_mecanico')   # mecánico a cargo (rol 4)
    id_estado    = request.form.get('id_estado', 1)
    # Se genera en el servidor para garantizar que no se repita
    numero_orden = modelo.obtener_siguiente_numero_orden()

    # Cliente: se busca por nombre; si no existe, se registra como nuevo (rol 3)
    cliente_nombre = (request.form.get('cliente') or '').strip()
    id_usuario = modelo.obtener_o_crear_cliente(cliente_nombre)

    # Vehículo: se busca por placa; si no existe, se registra como carro nuevo
    placa        = (request.form.get('placa') or '').strip().upper()
    marca_nombre = (request.form.get('marca') or '').strip()
    modelo_txt   = (request.form.get('modelo') or '').strip()

    vehiculo = modelo.obtener_vehiculo_por_placa(placa)
    if vehiculo:
        id_vehiculo = vehiculo['IDvehiculos']
    else:
        id_marca = modelo.obtener_o_crear_marca(marca_nombre)
        # color es NOT NULL en la BD; tipo por defecto 'Automóvil' (solo carros)
        id_vehiculo = modelo.crear_vehiculo(id_usuario, id_marca, placa, modelo_txt, None, '', 'Automóvil')

    ahora = datetime.now()
    # Fecha enviada desde el formulario; si no llega, usa la de hoy
    fecha_apertura = request.form.get('fecha_apertura') or ahora.strftime('%Y-%m-%d')
    hora_apertura  = ahora.strftime('%H:%M')

    # 1) Crea la orden (id_usuario = cliente)
    modelo.crear_orden(id_vehiculo, id_usuario, id_estado, numero_orden, hora_apertura, fecha_apertura)

    # 2) Crea la atención con el mecánico (rol 4) -> aparece en el Historial de Atenciones
    if id_mecanico:
        modelo.crear_atencion(id_vehiculo, id_mecanico, 4, fecha_apertura, None)

    return redirect(url_for('ordenes'))

@app.route('/ordenes/cerrar/<int:id>', methods=['POST'])
def cerrar_orden(id):
    id_estado    = request.form.get('id_estado', 2)
    fecha_cierre = request.form.get('fecha_cierre')
    total        = request.form.get('total', 0)
    modelo.actualizar_estado_orden(id, id_estado, fecha_cierre, total)
    return redirect(url_for('ordenes'))

@app.route('/ordenes/precio/<int:id>', methods=['POST'])
def precio_orden(id):
    # Solo registra el precio a cobrar (no cierra la orden; eso pasa al facturar)
    total = request.form.get('total', 0)
    modelo.actualizar_precio_orden(id, total)
    return redirect(url_for('ordenes'))

@app.route('/ordenes/finalizar/<int:id>')
def finalizar_orden(id):
    # Finalizar servicio: admins (1,2) y mecánico (4) — R004
    if session.get('id_rol') not in (1, 2, 4):
        return redirect(url_for('login'))
    modelo.finalizar_servicio(id, datetime.now().strftime('%Y-%m-%d'))
    return redirect(url_for('ordenes'))

@app.route('/ordenes/editar/<int:id>', methods=['POST'])
def editar_orden(id):
    # Solo Administrador (2) y Super_administrador (1) pueden editar
    if session.get('id_rol') not in (1, 2):
        return redirect(url_for('ordenes'))
    cliente    = request.form.get('cliente')
    placa      = (request.form.get('placa') or '').strip().upper()
    marca      = request.form.get('marca')
    modelo_txt = request.form.get('modelo')
    fecha      = request.form.get('fecha_apertura')
    modelo.actualizar_orden(id, cliente, placa, marca, modelo_txt, fecha)
    return redirect(url_for('ordenes'))

@app.route('/ordenes/eliminar/<int:id>')
def eliminar_orden(id):
    # Solo Administrador y Super_administrador pueden eliminar
    if session.get('id_rol') not in (1, 2):
        return redirect(url_for('ordenes'))
    modelo.eliminar_orden(id)
    return redirect(url_for('ordenes'))

@app.route('/ordenes/detalle/<int:id>')
def detalle_orden(id):
    detalle = modelo.obtener_detalle_orden(id)
    return jsonify(detalle)

@app.route('/ordenes/detalle/agregar', methods=['POST'])
def agregar_detalle():
    # Registrar repuestos/servicios: admins (1,2) y mecánico (4) — R004
    if session.get('id_rol') not in (1, 2, 4):
        return redirect(url_for('login'))
    id_orden         = request.form.get('id_orden')
    id_tipo_servicio = request.form.get('id_tipo_servicio')
    id_producto      = request.form.get('id_producto')
    cantidad         = request.form.get('cantidad')
    precio_unitario  = request.form.get('precio_unitario')
    modelo.agregar_detalle_orden(id_orden, id_tipo_servicio, id_producto, cantidad, precio_unitario)
    return redirect(url_for('ordenes'))

@app.route('/atenciones/agregar', methods=['POST'])
def agregar_atencion():
    id_vehiculo = request.form.get('id_vehiculo')
    id_usuario  = request.form.get('id_usuario')
    descripcion = request.form.get('descripcion')
    fecha       = request.form.get('fecha')
    modelo.crear_atencion(id_vehiculo, id_usuario, descripcion, fecha)
    return redirect(url_for('ordenes'))


# ═════════════════════════════════════════
#  FACTURACIÓN + MOVIMIENTOS
# ═════════════════════════════════════════

@app.route('/facturacion')
def facturacion():
    facturas         = modelo.obtener_facturas()
    lista_ord        = modelo.obtener_ordenes()
    metodos_pago     = modelo.obtener_metodos_pago()
    movimientos      = modelo.obtener_movimientos()
    productos        = modelo.obtener_productos()
    tipos_movimiento = modelo.obtener_tipos_movimiento()
    return render_template('Facturacion.html',
                           facturas=facturas,
                           ordenes=lista_ord,
                           metodos_pago=metodos_pago,
                           movimientos=movimientos,
                           productos=productos,
                           tipos_movimiento=tipos_movimiento,
                           ordenes_sin_factura=modelo.obtener_ordenes_sin_factura(),
                           siguiente_factura=modelo.obtener_siguiente_numero_factura(),
                           hoy=datetime.now().strftime('%Y-%m-%d'))

@app.route('/facturacion/agregar', methods=['POST'])
def agregar_factura():
    id_orden       = request.form.get('id_orden')
    id_metodo_pago = request.form.get('id_metodo_pago')
    fecha          = request.form.get('fecha')
    total          = request.form.get('total')
    # Evita facturar dos veces la misma orden (no duplicar)
    if id_orden and modelo.orden_tiene_factura(id_orden):
        return redirect(url_for('facturacion'))
    # El número de factura se genera en el servidor (consecutivo, no se repite)
    numero_factura = modelo.obtener_siguiente_numero_factura()
    modelo.crear_factura(id_orden, id_metodo_pago, numero_factura, fecha, total)
    # Al facturar, la orden ya terminó -> se cierra automáticamente (Finalizado)
    if id_orden:
        modelo.actualizar_estado_orden(id_orden, 2, fecha, total)
    return redirect(url_for('facturacion'))

@app.route('/movimientos/agregar', methods=['POST'])
def agregar_movimiento():
    id_producto        = request.form.get('id_producto')
    id_tipo_movimiento = request.form.get('id_tipo_movimiento')
    cantidad           = request.form.get('cantidad')
    fecha              = request.form.get('fecha')
    observacion        = request.form.get('observacion')
    modelo.registrar_movimiento(id_producto, id_tipo_movimiento, cantidad, fecha, observacion)
    return redirect(url_for('facturacion'))


# ═════════════════════════════════════════
#  PROVEEDORES
# ═════════════════════════════════════════

@app.route('/proveedores')
def proveedores():
    lista = modelo.obtener_proveedores()
    return render_template('proveedores.html', proveedores=lista)

@app.route('/proveedores/agregar', methods=['POST'])
def agregar_proveedor():
    nombre    = request.form.get('nombre')
    nit       = request.form.get('nit')
    telefono  = request.form.get('telefono')
    direccion = request.form.get('direccion')
    modelo.crear_proveedor(nombre, nit, telefono, direccion)
    return redirect(url_for('proveedores'))

@app.route('/proveedores/editar/<int:id>', methods=['POST'])
def editar_proveedor(id):
    nombre    = request.form.get('nombre')
    nit       = request.form.get('nit')
    telefono  = request.form.get('telefono')
    direccion = request.form.get('direccion')
    modelo.actualizar_proveedor(id, nombre, nit, telefono, direccion)
    return redirect(url_for('proveedores'))

@app.route('/proveedores/eliminar/<int:id>')
def eliminar_proveedor(id):
    modelo.eliminar_proveedor(id)
    return redirect(url_for('proveedores'))


# ═════════════════════════════════════════
#  ESTADÍSTICAS
# ═════════════════════════════════════════

MESES_ES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']


def _finanzas_con_extra(offset):
    d = modelo.obtener_finanzas_mes(offset)
    d['label'] = '%s %d' % (MESES_ES[d['mes'] - 1], d['anio'])
    d['neta'] = d['ganancias'] - d['gastos']
    d['pct_gastos'] = round(d['gastos'] / d['ganancias'] * 100) if d['ganancias'] else 0
    d['offset'] = offset
    d['esActual'] = offset == 0
    return d


@app.route('/api/finanzas-mes')
def api_finanzas_mes():
    if 'usuario' not in session:
        return jsonify({'error': 'no autorizado'}), 401
    try:
        offset = int(request.args.get('offset', 0))
    except (TypeError, ValueError):
        offset = 0
    offset = max(-600, min(0, offset))   # solo meses pasados y el actual
    return jsonify(_finanzas_con_extra(offset))


@app.route('/api/actividad-semanal')
def api_actividad_semanal():
    if 'usuario' not in session:
        return jsonify({'error': 'no autorizado'}), 401
    try:
        offset = int(request.args.get('offset', 0))
    except (TypeError, ValueError):
        offset = 0
    offset = max(-520, min(0, offset))   # solo semanas pasadas y la actual
    datos, lunes = modelo.obtener_ordenes_semana(offset)
    domingo = lunes + timedelta(days=6)
    meses = ['ene', 'feb', 'mar', 'abr', 'may', 'jun', 'jul', 'ago', 'sep', 'oct', 'nov', 'dic']
    if lunes.month == domingo.month:
        label = "%d–%d %s %d" % (lunes.day, domingo.day, meses[lunes.month - 1], lunes.year)
    else:
        label = "%d %s – %d %s %d" % (lunes.day, meses[lunes.month - 1], domingo.day, meses[domingo.month - 1], domingo.year)
    return jsonify({'datos': datos, 'label': label, 'offset': offset, 'esActual': offset == 0})


@app.route('/estadisticas')
def estadisticas():
    productos    = modelo.obtener_productos()
    lista_ord    = modelo.obtener_ordenes()
    facturas     = modelo.obtener_facturas()
    bajo_stock   = modelo.obtener_productos_bajo_stock()
    semana       = modelo.obtener_ordenes_semana_actual()
    mensual      = modelo.obtener_actividad_mensual_anio()   # órdenes por mes del año actual
    anio_actual  = datetime.now().year
    ordenes_pendientes = modelo.obtener_ordenes_sin_finalizar()
    finanzas     = _finanzas_con_extra(0)   # finanzas del mes actual

    # ── Datos reales para las tarjetas ──
    producto_top = modelo.obtener_producto_mas_vendido()   # {nombre_producto, total_ventas, ingresos}
    mecanicos    = modelo.obtener_mecanico_destacado()     # top 3 [{mecanico, total_ordenes}]
    ventas_cat   = modelo.obtener_ventas_por_categoria()

    # % del total y precio promedio del producto más vendido
    total_global = sum(c['total_ventas'] for c in ventas_cat) or 1
    if producto_top:
        producto_top['pct_total']   = round(producto_top['total_ventas'] / total_global * 100)
        producto_top['precio_prom'] = (producto_top['ingresos'] / producto_top['total_ventas']) if producto_top['total_ventas'] else 0

    # % de cada mecánico respecto al #1 (para las barritas)
    if mecanicos:
        tope = mecanicos[0]['total_ordenes'] or 1
        for m in mecanicos:
            m['pct'] = round(m['total_ordenes'] / tope * 100)

    return render_template('estadisticas.html',
                           productos=productos, ordenes=lista_ord, facturas=facturas,
                           bajo_stock=bajo_stock, semana=semana, mensual=mensual, anio_actual=anio_actual,
                           ordenes_pendientes=ordenes_pendientes, finanzas=finanzas,
                           producto_top=producto_top, mecanicos=mecanicos, ventas_cat=ventas_cat)

## ═════════════════════════════════════════
#  EXPORTAR EXCEL Y PDF
# ═════════════════════════════════════════
from flask import send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

@app.route('/exportar-excel')
def exportar_excel():
    wb = openpyxl.Workbook()

    # ── Hoja 1: Producto más vendido ──
    ws1 = wb.active
    ws1.title = "Producto Más Vendido"
    ws1.append(["Producto", "Ventas Totales", "Ingresos"])
    pm = modelo.obtener_producto_mas_vendido()
    if pm:
        ws1.append([pm['nombre_producto'], pm['total_ventas'], float(pm['ingresos'])])

    # ── Hoja 2: Ventas por Categoría ──
    ws2 = wb.create_sheet("Categorías")
    ws2.append(["Categoría", "Ventas", "Ingresos"])
    for c in modelo.obtener_ventas_por_categoria():
        ws2.append([c['categoria'], c['total_ventas'], float(c['ingresos'])])

    # ── Hoja 3: Actividad Mensual ──
    ws3 = wb.create_sheet("Actividad Mensual")
    meses = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
             7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}
    ws3.append(["Mes", "Órdenes"])
    for m in modelo.obtener_actividad_mensual():
        ws3.append([meses.get(m['mes'], m['mes']), m['total_ordenes']])

    # ── Hoja 4: Ganancias y Gastos ──
    ws4 = wb.create_sheet("Ganancias y Gastos")
    ws4.append(["Concepto", "Total"])
    ws4.append(["Total Ganancias", float(modelo.obtener_total_ganancias())])
    ws4.append(["Total Gastos",    float(modelo.obtener_total_gastos())])

    # ── Hoja 5: Mecánicos Destacados ──
    ws5 = wb.create_sheet("Mecánicos")
    ws5.append(["Mecánico", "Órdenes Atendidas"])
    for m in modelo.obtener_mecanico_destacado():
        ws5.append([m['mecanico'], m['total_ordenes']])

    # ── Hoja 6: Órdenes ──
    ws6 = wb.create_sheet("Órdenes")
    ws6.append(["N° Orden", "Cliente", "Vehículo", "Fecha Apertura", "Fecha Cierre", "Estado", "Total"])
    for o in modelo.obtener_ordenes():
        ws6.append([o['numero_orden'], o['cliente'], o['placa'],
                    str(o['fecha_apertura']), str(o['fecha_cierre']),
                    o['estado'], float(o['total'] or 0)])

    # ── Hoja 7: Facturas ──
    ws7 = wb.create_sheet("Facturas")
    ws7.append(["N° Factura", "Orden", "Fecha", "Método Pago", "Total"])
    for f in modelo.obtener_facturas():
        ws7.append([f['numero_factura'], f['numero_orden'],
                    str(f['fecha']), f['metodo_pago'], float(f['total'])])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="reporte_taller_el_costa.xlsx", as_attachment=True)


@app.route('/exportar-pdf')
def exportar_pdf():
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    naranja = colors.HexColor('#f97316')

    # Título
    elements.append(Paragraph("Reporte General — Taller El Costa", styles['Title']))
    elements.append(Spacer(1, 12))

    # ── Producto más vendido ──
    elements.append(Paragraph("Producto Más Vendido", styles['Heading2']))
    pm = modelo.obtener_producto_mas_vendido()
    if pm:
        data = [["Producto", "Ventas Totales", "Ingresos"],
                [pm['nombre_producto'], pm['total_ventas'], f"${float(pm['ingresos']):,.0f}"]]
        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), naranja),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f5f5f5')]),
        ]))
        elements.append(t)
    elements.append(Spacer(1, 12))

    # ── Ventas por Categoría ──
    elements.append(Paragraph("Ventas por Categoría", styles['Heading2']))
    data = [["Categoría", "Ventas", "Ingresos"]]
    for c in modelo.obtener_ventas_por_categoria():
        data.append([c['categoria'], c['total_ventas'], f"${float(c['ingresos']):,.0f}"])
    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), naranja),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f5f5f5'), colors.white]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 12))

    # ── Ganancias y Gastos ──
    elements.append(Paragraph("Ganancias y Gastos", styles['Heading2']))
    data = [["Concepto", "Total"],
            ["Total Ganancias", f"${float(modelo.obtener_total_ganancias()):,.0f}"],
            ["Total Gastos",    f"${float(modelo.obtener_total_gastos()):,.0f}"]]
    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), naranja),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f5f5f5'), colors.white]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 12))

    # ── Mecánicos Destacados ──
    elements.append(Paragraph("Top Mecánicos", styles['Heading2']))
    data = [["Mecánico", "Órdenes Atendidas"]]
    for m in modelo.obtener_mecanico_destacado():
        data.append([m['mecanico'], m['total_ordenes']])
    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), naranja),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f5f5f5'), colors.white]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 12))

    # ── Órdenes ──
    elements.append(Paragraph("Órdenes de Servicio", styles['Heading2']))
    data = [["N° Orden", "Cliente", "Placa", "Fecha", "Estado", "Total"]]
    for o in modelo.obtener_ordenes():
        data.append([o['numero_orden'], o['cliente'], o['placa'],
                     str(o['fecha_apertura']), o['estado'],
                     f"${float(o['total'] or 0):,.0f}"])
    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), naranja),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f5f5f5'), colors.white]),
        ('FONTSIZE', (0,0), (-1,-1), 8),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, download_name="reporte_taller_el_costa.pdf", as_attachment=True)